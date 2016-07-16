#! python3

import os
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Style

print(os.getcwd())

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))

sheetNames = wb.get_sheet_names()
print(sheetNames)

sheet = wb.get_sheet_by_name('Sheet3')  # to obtain a Worksheet object
print(sheet)

print(sheet.title)

##activeSheet = wb.get_active_sheet()  # this is a deprecated call - use .active property
##print(activeSheet)

activeSheet = wb.active
print("active sheet is {}".format(activeSheet))

# once you have a Worksheet object, you can access a Cell object by its name

sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'].value)   # openpyxl converts to datetime values instead of strings

c = sheet['B1']
print(c.value)

print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)

print('Cell ' + c.coordinate + ' is ' + c.value)

print(sheet['C1'].value)

# print cell location and value
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

# to print the fruit in the odd numbered rows
for i in range(1,8,2):
    print(i, sheet.cell(row=i, column=2).value)

# You can determine the size of the sheet with the Worksheet's object's get_highest_row()
# and get_highest_column() methods

##print(sheet.get_highest_row())    # deprecated call - use max_row property
##print(sheet.get_highest_column()) # deprecated call - use max_column property

print(sheet.max_row)
print(sheet.max_column)  # returns integer not letter(s)

print(openpyxl.cell.get_column_letter(1))

# after using "from openpyxl.cell import get_column_letter, column_index_from_string" can use the
# methods directly as follows

print(get_column_letter(1))   # A
print(get_column_letter(27))  # AA
print(get_column_letter(900)) # AHP

print(get_column_letter(sheet.max_column))  # C

# to convert column index from the letter string
print(column_index_from_string('A'))  # 1
print(column_index_from_string('AA')) # 27

#----------------------------------------------------------------------------------
# You can slice Worksheet objects to get all the Cell objects in a row, column, or
# retangular area of the spreadsheet.  Then you can loop over all the cells in the slice
# see example in the following code

# sheet = wb.get_sheet_by_name('Sheet')   # for reference - already defined earlier in code
print(tuple(sheet['A1':'C3']))
'''outputs the following:

((<Cell Sheet1.A1>, <Cell Sheet1.B1>, <Cell Sheet1.C1>), (<Cell Sheet1.A2>, <Cell Sheet1.B2>,
<Cell Sheet1.C2>), (<Cell Sheet1.A3>, <Cell Sheet1.B3>, <Cell Sheet1.C3>))

'''

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---END OF ROW---')
'''outputs the following:
A1 2015-04-05 13:34:02
B1 Apples
C1 73
---END OF ROW---
A2 2015-04-05 03:41:23
B2 Cherries
C2 85
---END OF ROW---
A3 2015-04-06 12:46:51
B3 Pears
C3 14
---END OF ROW---

'''
#----------------------------------------------------------------------------------
# To access the values of cells in a particular row or column, you can also use a
# Worksheet object's rows and columns attribute.
# see example in the following code

print(sheet.columns[1])

'''outputs the following:
(<Cell Sheet1.B1>, <Cell Sheet1.B2>, <Cell Sheet1.B3>, <Cell Sheet1.B4>, <Cell Sheet1.B5>,
<Cell Sheet1.B6>, <Cell Sheet1.B7>)
'''
for cellObj in sheet.columns[1]:
    print(cellObj.value)

''' outputs the following:
Apples
Cherries
Pears
Oranges
Apples
Bananas
Strawberries
'''

for cellObj in sheet.rows[1]:
    print(cellObj.value)

''' outputs the following:
2015-04-05 03:41:23
Cherries
85
'''
#----------------------------------------------------------------------------------      
# Creating and Saving Excel Documents

wbk = openpyxl.Workbook()   # create new, blank Workbook object

print(wbk.get_sheet_names())  # Sheet
sheet = wbk.active
print('Sheet title = ', sheet.title)

sheet.title = "Spam Bacon Eggs Sheet"
print('Sheet title is now = ', sheet.title)

print(wbk.get_sheet_names())

wbk.save('example_2.xlsx')

#-------------------------------------------------------------------------------
# Creating and Removing Sheets

print(wbk.get_sheet_names())

sheet.title = 'Sheet'

wbk.create_sheet()
print(wbk.get_sheet_names())  # ['Sheet', 'Sheet1']

# can set parameters when creating new sheets
wbk.create_sheet(index=0, title='First Sheet')
print(wbk.get_sheet_names())  # ['First Sheet', 'Sheet', 'Sheet1']

wbk.create_sheet(index=2, title='Middle Sheet')
print(wbk.get_sheet_names())  # ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']

# the remove_sheet() takes a Workbook object, not a string, must do the following
wbk.remove_sheet(wbk.get_sheet_by_name('Middle Sheet'))
wbk.remove_sheet(wbk.get_sheet_by_name('Sheet1'))
print(wbk.get_sheet_names())  # ['First Sheet', 'Sheet']

#-------------------------------------------------------------------------------
# Writing Values to Cells

sheet['A1'] = 'Hello world!'
print('cell A1 = ', sheet['A1'].value)

#-------------------------------------------------------------------------------
# Setting the Font Styles of Cells
# To customize font styles in cells it is important to import the Font() and Style()
# use 'from openpyxl.styles import Font, Style'

wbk2 = openpyxl.Workbook()
sheet = wbk2.get_sheet_by_name('Sheet')

italic24Font = Font(size=24, italic=True)

'''
styleObj = Style(font=italic24Font) #1
sheet['A1'].style=styleObj          #2

Warning (from warnings module):
  File "C:\Python34\lib\site-packages\openpyxl\styles\styleable.py", line 189
    warn("Use formatting objects such as font directly")
UserWarning: Use formatting objects such as font directly


'''

sheet['A1'].font = italic24Font #used to replace above #1 & #2 to eliminate warning
sheet['A1'] = 'Hello world!'
wbk2.save('styled.xlsx')

#-------------------------------------------------------------------------------
# Formulas

wbk3 = openpyxl.Workbook()
sheet = wbk3.active    # get active worksheet

sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wbk3.save('writeFormula.xlsx')

wbFormulas = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wbFormulas.active
print('data_only = True --> ', sheet['A3'].value)

wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.active
print('data_only = False (default) --> ', sheet['A3'].value)

#-------------------------------------------------------------------------------
# Adjusting Rows and Columns

sheet['B1'] = 'Tall row'
sheet['C2'] = 'Wide column'
sheet.row_dimensions[1].height = 70      # set to int or float btwn 0 and 409, default = 12.75
sheet.column_dimensions['C'].width = 20  # set to int or float btwn 0 and 255, default = 8.43 chars
                                         # to hide a column or row, set width or height to 0

wbFormulas.save('dimensions.xlsx')
print('Done adjusting rows, columns.')

#-------------------------------------------------------------------------------
# Merging and Unmerging Cells

sheet.merge_cells('A7:D9')
sheet['A7'] = 'Twelve cells merged together'

sheet.merge_cells('C11:D11')
sheet['C11'] = 'Two merged cells'

wbFormulas.save('merged.xlsx')
print('Done merging cells')

# unmerge the previous cells
sheet.unmerge_cells('A7:D9')
sheet.unmerge_cells('C11:D11')
wbFormulas.save('unmerged.xlsx')
print("Done unmerging.")

#-------------------------------------------------------------------------------
# Freeze Panes

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')
print('Done with freeze pane.')

      

































